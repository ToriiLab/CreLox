
#Plots cotyledon and sector outlines as polygons, and plots stomata as point distributions.
#Calculates stomatal distances from sectors and histograms the results.
#Generates N random point distributions within each cotyledon and repeats this process.

#This script is capable of analyzing data from more than one file at a time if
#you supply a list of file directories.

import time #If you want to track the length of the script
start_time = time.time()

import numpy as np;
import csv 
import openpyxl;
import matplotlib.mlab as mlab;
import matplotlib.pyplot as pyplot;
import random;
import math;
import scipy.spatial as sps;
import shapely.geometry as shg;

'''
Begin section for defining necessary functions.
'''

# Takes a workbook sheet and a point count and returns an Nx2 numpy array of point values
def recordSheetCoordinates(point_sheet, point_count):
    points = np.zeros((point_count, 3), dtype = np.float);
    for i in range(0, point_count):
        points[i, 0] = point_sheet.cell(row = i + 3, column = 1).value;
        points[i, 1] = point_sheet.cell(row = i + 3, column = 2).value;
    
    # Calculate angle of each point from center (x_mean, y_mean) and sort points by angle 
    x_mean = np.mean(points[:, 0]);
    y_mean = np.mean(points[:, 1]);
    for j in range(0, point_count):
        points[j, 2] = math.atan2(points[j, 1] - y_mean, points[j, 0] - x_mean);
    points = points[points[:,2].argsort()];
    
    # New array with only the x-value and y-value columns
    points2 = np.delete(points,np.s_[-1:],1);
    return points2;

# Check if the given list of points are inside the polygon specified by the list of
# polygon_points. Returns the indices of the points_list whose points are inside the
# given polygon
def checkPointsInPolygon(polygon_list, points_list):
    inside = mlab.inside_poly(points_list, polygon_list);
    return inside;

# Takes the number of random points to be generated, the maximum x-coordinate value
# of the cotyledon outline, and the maximum y-coordinate value of the cotyledon
# outline and returns an Nx2 numpy array of random points.
def generateRandomPoints(number_of_points, x_max, y_max):
    random_points = np.zeros((number_of_points, 2), dtype = np.float);
    for i in range(0, number_of_points):
        random_points[i, 0] = random.uniform(0, x_max);
        random_points[i, 1] = random.uniform(0, y_max); 
    return random_points;
    
#Creates the polygon defined by the sector outline and computes the nearest distance
#between each individual point of your sapmle distribution and the sector.
#points_list is the point distribution, and outline is the sector. 
#Returns a nx2 array where n is the length of points_list, which must also be an nx2 array.

    
def computeSectorDistance(points_list,outline):
    point_count = len(points_list);
    poly = shg.Polygon(outline);
    distance_list = []
    for i in range(0,point_count):
        point = shg.Point(points_list[i]);
        distance_list.append(poly.distance(point));
    distance_list = np.asarray(distance_list);
    distances = distance_list[distance_list!=0];
    return distances;


'''
End of function definition section.
'''    
    

'''
Begin of procedural section.
'''

# Reads in the Excel File and gets the worksheet names
# MAKE SURE EXCEL WORKBOOK FILE IS IN FORMAT .xlsx, NOT .xls, OTHERWISE
# THE CODE WILL NOT RUN!

min_area = 0; #If you want to place an upper and lower limit on size of sectors to analyze
max_area = 10000000;

directory = [];
save = [];
phenotype = [];

Tota_Sector_Data = [];

Sector_Stomata = [];
Sector_Random = [];
Sector_Ratio = [];

Area_Data = [];

#Open the file list that contains the pathway for each dataset meant to be analyzed.

#filelist.csv should be the name of a list of pathways to each dataset file you want analyzed.
#The first column should be the pathway to each dataset file, the second column is a savename, and the third column is phenotype.

#Each dataset should correspond to one cotyledon,
#which may have multiple sectors in separate worksheets within the xlsx file.

with open("filelist.csv", 'rU') as f:  #see above for what you should put in for filelist.csv 
    text = csv.reader(f);
    next(text, None);  # skip the headers
    
    for row in text: 
        directory.append(row[0]);
        save.append(row[1]);
        phenotype.append(row[2]);
        
#For loops that iterates over each individual dataset that the file list specifies.

#Your xlsx file should have a worksheet should have a worksheet named Stomatal Postions,
#a worksheet named Cotyledon Outline, and worksheets named Sector 1 Outline, Sector 2 Outline, etc.
    
#Each dataset should correspond to one cotyledon, which may have multiple sectors.
        
for i in range(0,len(directory)):


    #Loads the ith dataset in the filelist.

    crelox_data_file_name = directory[i];
    title = save[i];
    leaf_type = phenotype[i];

    crelox_wb = openpyxl.load_workbook(crelox_data_file_name);
    crelox_wb_sheet_names = crelox_wb.get_sheet_names();
    
    # Reads in the data from the worksheets
    stomata_sheet = crelox_wb.get_sheet_by_name('Stomatal Positions');
    cotyledon_sheet = crelox_wb.get_sheet_by_name('Cotyledon Outline');
    
    stomata_count = stomata_sheet.max_row - 2;
    cotyledon_point_count = cotyledon_sheet.max_row - 2;
    
    stomata_points = recordSheetCoordinates(stomata_sheet, stomata_count);
    cotyledon_points = recordSheetCoordinates(cotyledon_sheet, cotyledon_point_count);
    cot_x_max = cotyledon_points[:,0].max();
    cot_y_max = cotyledon_points[:,1].max();
    
    #Counts the number of stomata inside the cotyledon and removes any coordinates
    #that are somehow outside of the cotyledon boundary
    
    stomata_points_indices_inside = checkPointsInPolygon(cotyledon_points, stomata_points);
    number_inside = len(stomata_points_indices_inside);
    stomata_points_inside = np.zeros((number_inside, 2), dtype=np.int);
    for i in range(0, number_inside):
        stomata_points_inside[i] = stomata_points[stomata_points_indices_inside[i]];
        
    stomata_points = stomata_points_inside
    

    
    # Gets the number of sector outline worksheets. Relies on the fact that the sector
    # worksheets come sequentially immediately following the cotyledon outline worksheet
    
    #Sector Outlines shouldn't go up to 15, but list has been written to a high number just in case.    

    sector_names = ['Sector 1 Outline', 'Sector 2 Outline', 'Sector 3 Outline', 'Sector 4 Outline', 'Sector 5 Outline', 'Sector 6 Outline', 'Sector 7 Outline', 'Sector 8 Outline', 'Sector 9 Outline', 'Sector 10 Outline', 'Sector 11 Outline', 'Sector 12 Outline', 'Sector 13 Outline', 'Sector 14 Outline', 'Sector 15 Outline'];
    cot_sheet_index = crelox_wb_sheet_names.index('Cotyledon Outline');
    number_of_sectors = len(crelox_wb_sheet_names) - (cot_sheet_index + 1);
    
    sector_list = ['Sector 1', 'Sector 2', 'Sector 3', 'Sector 4', 'Sector 5', 'Sector 6', 'Sector 7', 'Sector 8', 'Sector 9', 'Sector 10', 'Sector 11', 'Sector 12', 'Sector 13', 'Sector 14', 'Sector 15'];

    #Generate N random distributions for correlation function calculations

    Ntrials = 1000;
        
    random_sets = [];
    
    #First generate an excess of random points within a rectangular boundary,
    #then keep the first N inside of the cotyledon, where N is the number of
    #real stomata inside the cotyledon. Repeat this process for the specified number of trials.
    
        
    for i in range(0,Ntrials):
        
        all_rand_points2 = generateRandomPoints(stomata_count*3, cot_x_max, cot_y_max);
        rand_points_indices_inside2 = checkPointsInPolygon(cotyledon_points, all_rand_points2);
        number_inside2 = len(rand_points_indices_inside2);
        rand_points_inside2 = np.zeros((number_inside2, 2), dtype=np.int);
        for i in range(0, number_inside2):
            rand_points_inside2[i] = all_rand_points2[rand_points_indices_inside2[i]];
        random_dist = rand_points_inside2[:stomata_count];    
        N_number_inside2 = len(random_dist);    
        
        random_sets.append(random_dist);
    
    sector_distances = [];
    random_distances = [];
    
    #Load each sector on the cotyledon and calculate distances from the stomata to the sector
    #and the distances from each random point distribution to the sector        
    
    for i in range(0, number_of_sectors):
        sector_sheet = crelox_wb.get_sheet_by_name(sector_names[i]);
        sector_point_count = sector_sheet.max_row - 2;
        sector_points = recordSheetCoordinates(sector_sheet, sector_point_count);
        
        poly = shg.Polygon(sector_points);

        if min_area <= poly.area <= max_area: #area filter on sectors if desired
        
            Area_Data.append(poly.area);
        
            distances = computeSectorDistance(stomata_points,sector_points);
            sector_distances.append(distances);
            random_list = [];
            for i in range(0,Ntrials):    
                rand_dist = computeSectorDistance(random_sets[i],sector_points);
                b = rand_dist.tolist();
                random_list = random_list + b;  
            random_distances.append(random_list);
            
    #Define logarithmic distance bins for histogramming
    
    distance_fun = np.logspace(np.log10(15),np.log10(2500),15);
    distance_bins = np.insert(distance_fun,0,0);
    
    #Histogram the stomata and random point distances
    
    #Random distances are histogrammed in aggregate, rather than individually, as we only
    #care about approximating the expected value.

    
    sector_data = np.empty((number_of_sectors,len(distance_bins)-1));
    random_data = np.empty((number_of_sectors,len(distance_bins)-1));
    ratio_data = np.empty((number_of_sectors,len(distance_bins)-1));
    
    for i in range(0,len(sector_distances)):
        
        sector_data,bin_data,etc = pyplot.hist(sector_distances[i], bins = distance_bins, alpha = 0.5, label = 'Stomata-Sector distances', color = 'b',histtype='step');    
        random_data,bin_data,etc = pyplot.hist(random_distances[i], bins = distance_bins, alpha = 0.5, label = 'Random-Sector distances', color = 'r',histtype='step');

        Sector_Stomata.append(sector_data);
        Sector_Random.append(random_data);
    
        ratio_data = (sector_data/float(len(sector_distances)))/(random_data/float(len(random_distances)))-1;
        Sector_Ratio.append(ratio_data);
    
#Save the histogram counts and sector area data

#Sector_Data is a 3xnxN array

#The first dimension is stomata distance histograms, random point histograms, and normalized ratio data.

#The second dimension is the length of the number of distance bins n

#The third dimension is the total number of distance histograms N
    
areas = np.asarray(Area_Data);

Total_Sector_Data = [Sector_Stomata,Sector_Random,Sector_Ratio];
Sector_Data = np.asarray(Total_Sector_Data);


np.save('filename here 1.npy',areas); #Filename for area data
np.save('filename here 2.npy',Sector_Data); #File name for histogram data


print("--- %s seconds ---" % (time.time() - start_time)); #If you want to know how long it took
'''
End of procedural section.
'''