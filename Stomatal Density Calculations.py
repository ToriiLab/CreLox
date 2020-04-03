#Plots cotyledon and sector outlines as polygons, and plots stomata as point distributions.
#Calculates the density of stomata within defined regions within the sector or cotyledons.
#Exports the data into npy files.

#This script is capable of analyzing data from more than one file at a time if
#you supply a list of file directories.

import time #Only here if you need to know how long the script is running
start_time = time.time()

import numpy as np;
import csv 
import openpyxl;
import matplotlib.mlab as mlab;
import matplotlib.pyplot as pyplot;
import random;
import math;
import scipy.spatial as sps;
import shapely.geometry as shg;
from shapely.ops import cascaded_union

#
'''
Begin section for defining necessary functions.
'''

# Takes a workbook sheet and a point count and returns an Nx2 numpy array of point values.


def recordSheetCoordinates(point_sheet, point_count):
    points = np.zeros((point_count, 3), dtype = np.float);
    for i in range(0, point_count):
        points[i, 0] = point_sheet.cell(row = i + 3, column = 1).value;
        points[i, 1] = point_sheet.cell(row = i + 3, column = 2).value;
    
    # Calculate angle of each point from center (x_mean, y_mean) and sort points by angle 
    x_mean = np.mean(points[:, 0]);
    y_mean = np.mean(points[:, 1]);
    for j in range(0, point_count):
        points[j, 2] = math.atan2(points[j, 1] - y_mean, points[j, 0] - x_mean);
    points = points[points[:,2].argsort()];
    
    # New array with only the x-value and y-value columns
    points2 = np.delete(points,np.s_[-1:],1);
    return points2;

# Check if the given list of points are inside the polygon specified by the list of
# polygon_points. Returns the indices of the points_list whose points are inside the
# given polygon
def checkPointsInPolygon(polygon_list, points_list):
    inside = mlab.inside_poly(points_list, polygon_list);
    return inside;

# Takes the number of random points to be generated, the maximum x-coordinate value
# of the cotyledon outline, and the maximum y-coordinate value of the cotyledon
# outline and returns an Nx2 numpy array of random points.
def generateRandomPoints(number_of_points, x_max, y_max):
    random_points = np.zeros((number_of_points, 2), dtype = np.float);
    for i in range(0, number_of_points):
        random_points[i, 0] = random.uniform(0, x_max);
        random_points[i, 1] = random.uniform(0, y_max); 
    return random_points;

    
#Creates the polygon defined by the sector outline and computes the nearest distance
#between each individual point of your sapmle distribution and the sector.
#points_list is the point distribution, and outline is the sector. 
#Returns a nx2 array where n is the length of points_list, which must also be an nx2 array.
    
def computeSectorDistance(points_list,outline):
    point_count = len(points_list);
    poly = shg.Polygon(outline);
    distance_list = []
    for i in range(0,point_count):
        point = shg.Point(points_list[i]);
        distance_list.append(poly.distance(point));
    distance_list = np.asarray(distance_list);
    distances = distance_list[distance_list!=0];
    return distances;
    
#Creates a virtual sector, given a sector outline
#and a radius to extend the range of the sector by.
#Used to check stomatal density at an extended range away from the real sector.
    
def SectorRange(sector_points, radius):
    x_mean = np.mean(sector_points[:,0]);
    y_mean = np.mean(sector_points[:,1]);
    direct_vectors = sector_points - (x_mean,y_mean);
    unit_vectors = np.zeros(np.shape(direct_vectors));
    for i in range(0,len(unit_vectors)):
        unit_vectors[i,0] = direct_vectors[i,0]/np.sqrt(np.dot(direct_vectors[i],direct_vectors[i]));
        unit_vectors[i,1] = direct_vectors[i,1]/np.sqrt(np.dot(direct_vectors[i],direct_vectors[i]));
    new_sector = sector_points + radius * unit_vectors;
    
    return new_sector;


'''
End of function definition section.
'''    
    

'''
Begin of procedural section.
'''

# Reads in the Excel File and gets the worksheet names
# MAKE SURE EXCEL WORKBOOK FILE IS IN FORMAT .xlsx, NOT .xls, OTHERWISE
# THE CODE WILL NOT RUN!

directory = [];
save = [];
phenotype = [];
in_range_densities = [];
outside_range_densities = [];
sector_densities = []

#Open the file list that contains the pathway for each dataset meant to be analyzed.

#filelist.csv should be the name of a list of pathways to each dataset file you want analyzed.
#The first column should be the pathway to each dataset file, the second column is a savename, and the third column is phenotype.

#Each dataset should correspond to one cotyledon,
#which may have multiple sectors in separate worksheets within the xlsx file.

with open("filelist.csv", 'rU') as f:  #see above for what you should put in for filelist.csv 
    text = csv.reader(f);
    next(text, None);  # skip the headers
    
    for row in text: 
        directory.append(row[0]);
        save.append(row[1]);
        phenotype.append(row[2]);

sectors_overall = [];
inside_overall = [];
outside_overall = [];

#For loops that iterates over each individual dataset that the file list specifies.

#Your xlsx file should have a worksheet should have a worksheet named Stomatal Postions,
#a worksheet named Cotyledon Outline, and worksheets named Sector 1 Outline, Sector 2 Outline, etc.
    
#Each dataset should correspond to one cotyledon, which may have multiple sectors.
    
for i in range(0,len(directory)):
    
    #Loads the ith dataset in the filelist.

    crelox_data_file_name = directory[i];
    title = save[i];
    leaf_type = phenotype[i];

    crelox_wb = openpyxl.load_workbook(crelox_data_file_name);
    crelox_wb_sheet_names = crelox_wb.get_sheet_names();
    
    # Reads in the stomata and cotyledon data from the worksheets
    stomata_sheet = crelox_wb.get_sheet_by_name('Stomatal Positions');
    cotyledon_sheet = crelox_wb.get_sheet_by_name('Cotyledon Outline');
    
    stomata_count = stomata_sheet.max_row - 2;
    cotyledon_point_count = cotyledon_sheet.max_row - 2;
    
    stomata_points = recordSheetCoordinates(stomata_sheet, stomata_count);
    cotyledon_points = recordSheetCoordinates(cotyledon_sheet, cotyledon_point_count);
    cot_x_max = cotyledon_points[:,0].max();
    cot_y_max = cotyledon_points[:,1].max();

    
    # Gets the number of sector outline worksheets. Relies on the fact that the sector
    # worksheets come sequentially immediately following the cotyledon outline worksheet

    #Sector Outlines shouldn't go up to 15, but list has been written to a high number just in case.    
    
    sector_names = ['Sector 1 Outline', 'Sector 2 Outline', 'Sector 3 Outline', 'Sector 4 Outline', 'Sector 5 Outline', 'Sector 6 Outline', 'Sector 7 Outline', 'Sector 8 Outline', 'Sector 9 Outline', 'Sector 10 Outline', 'Sector 11 Outline', 'Sector 12 Outline', 'Sector 13 Outline', 'Sector 14 Outline', 'Sector 15 Outline'];
    cot_sheet_index = crelox_wb_sheet_names.index('Cotyledon Outline');
    number_of_sectors = len(crelox_wb_sheet_names) - (cot_sheet_index + 1);
    
    sector_list = ['Sector 1', 'Sector 2', 'Sector 3', 'Sector 4', 'Sector 5', 'Sector 6', 'Sector 7', 'Sector 8', 'Sector 9', 'Sector 10', 'Sector 11', 'Sector 12', 'Sector 13', 'Sector 14', 'Sector 15'];
    
    # Gets the correct worksheets for sector outlines based on number_of_sectors
    #Also calculates area of sectors and number of stomata
    
    colors = ['g','c','k','r','y','r','m','g','c','k','y','r','m','g','c','k','y']
    
    polygon_list = []; 
    
    stomata_inside_sectors = 0
    total_area_of_sectors = 0
    
    for i in range(0, number_of_sectors):
        
        #Extract sector coordinates
        
        sector_sheet = crelox_wb.get_sheet_by_name(sector_names[i]);
        sector_point_count = sector_sheet.max_row - 2;
        sector_points = recordSheetCoordinates(sector_sheet, sector_point_count);
        pyplot.plot(sector_points[:,0], sector_points[:,1], color=colors[i],label=sector_list[i]);

        #Creates virtual sector based off desired range.

        range_sector = SectorRange(sector_points, 100);
        pyplot.plot(range_sector[:,0], range_sector[:,1], color=colors[i],label=sector_list[i]);
    
        #Stores sectors and virual sectors as polygons.    
        
        sector_poly = shg.Polygon(sector_points);
        poly = shg.Polygon(range_sector);
        polygon_list.append(poly);
        
        #Aggregates number of stomata inside sectors for a given cotyledon
        
        stomata_inside = len(checkPointsInPolygon(sector_points,stomata_points));
        stomata_inside_sectors = stomata_inside + stomata_inside_sectors;
        
        #Aggregates area of sectors for a given cotyledon
        
        total_area_of_sectors = total_area_of_sectors + sector_poly.area;
        
    #Calculate the total area of the cotyledon for virtual sectors,
    #accounting for potential overlap due to the extended ranges
    #by defining the Union of the polygons

    polygon_list.append(shg.Polygon(cotyledon_points));   
    
    inter = cascaded_union([polygon_list[number_of_sectors].intersection(polygon_list[i]) for i in range(0,number_of_sectors)]);

    #Find stomata within the cotyledon but outside of the virtual sectors,
    #i.e. out_of_range.
    #Also find stomata within the virtual sectors but not within the sectors themselves, i.e.
    #in_range

    out_of_range = 0;
    
    
    for i in range(0,len(stomata_points)):
        point = shg.Point(stomata_points[i])
        if inter.contains(point) == False:  #Checks if point is inside the union of virtual sectors
            out_of_range = out_of_range+1;
                        
    in_range = len(stomata_points) - out_of_range - stomata_inside_sectors;
    
    #Now that we've counted stomata in the defined regions, now we calculate area for those regions.    
    
    total_area = polygon_list[number_of_sectors].area;
    
    sector_range_area = inter.area - total_area_of_sectors; #in_range has to exclude the sector itself
    outside_range_area = total_area - inter.area;
    
    #Calculate the densities in the defined regions
    
    in_range_density = in_range/sector_range_area;
    outside_range_density = out_of_range/outside_range_area;
    
    total_sector_density = stomata_inside_sectors / total_area_of_sectors;
    
    #Store the densities into arrays
    
    in_range_densities.append(in_range_density);
    outside_range_densities.append(outside_range_density);
    sector_densities.append(total_sector_density);
    
    sectors_overall.append(stomata_inside_sectors);
    inside_overall.append(in_range);
    outside_overall.append(out_of_range);
    
    #Repeat for loop for other cotyledons in filelist

#Save three npy files: the stomatal densities in the three defined regions of
#sectors, the extended virtual range of the sectors, and the remainder of the cotyledon

#The ith density in each file corresponds to the ith cotyledon in your filelist.
    
inside = np.asarray(in_range_densities);
outside = np.asarray(outside_range_densities);
sector = np.asarray(sector_densities);

np.save('virtual sector filename.npy',inside);
np.save('rest of cotyledon filename.npy',outside);
np.save('real sector filename.npy',sector); 



'''
End of procedural section.
'''